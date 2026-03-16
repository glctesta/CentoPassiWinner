"""
Centopassi Route Planner - XLSX Waypoint Loader
Parses waypoint data from Excel files for seasonal updates.

Supports flexible column detection:
  - Columns can be in any order
  - Auto-detects common column names (name, lat, lon, elevation, etc.)
  - Identifies Golden Points and unpaved roads from name/description
"""
import os
import re
from models import Waypoint
from config import UNPAVED_ROAD_KEYWORDS, FORBIDDEN_ROAD_KEYWORDS
from gpx_parser import classify_road_type

# Column name aliases (case-insensitive matching)
COLUMN_ALIASES = {
    'name': ['name', 'nome', 'waypoint', 'wp', 'numero', 'number', 'id', 'n°', 'n.'],
    'lat': ['lat', 'latitude', 'latitudine', 'y'],
    'lon': ['lon', 'lng', 'longitude', 'longitudine', 'x'],
    'elevation': ['elevation', 'ele', 'altitude', 'quota', 'altitudine', 'alt', 'elev'],
    'description': ['description', 'desc', 'descrizione', 'note', 'notes', 'dettagli',
                     'road', 'strada', 'tipo', 'type', 'road_type'],
    'city': ['city', 'città', 'citta', 'comune', 'locality', 'località', 'localita'],
    'province': ['province', 'provincia', 'prov', 'state', 'regione', 'region'],
    'golden': ['gp', 'golden', 'golden_point', 'goldenpoint', 'gold', 'tipo_wp'],
}


def _match_column(header: str, alias_group: str) -> bool:
    """Check if a column header matches an alias group."""
    header_lower = header.strip().lower().replace(' ', '_')
    for alias in COLUMN_ALIASES.get(alias_group, []):
        if alias == header_lower or header_lower.startswith(alias):
            return True
    return False


def _detect_columns(headers: list[str]) -> dict[str, int]:
    """
    Auto-detect column mapping from headers.
    Returns {field_name: column_index}.
    """
    mapping = {}
    used_indices = set()
    
    # Priority order: name > lat > lon > elevation > description > city > province > golden
    for field in ['name', 'lat', 'lon', 'elevation', 'description', 'city', 'province', 'golden']:
        for idx, header in enumerate(headers):
            if idx in used_indices:
                continue
            if _match_column(header, field):
                mapping[field] = idx
                used_indices.add(idx)
                break
    
    return mapping


def parse_xlsx(filepath: str, sheet_name: str = None) -> list[Waypoint]:
    """
    Parse an XLSX file and return list of Waypoint objects.
    
    The file can have columns in any order. Column names are matched
    against known aliases. At minimum, 'name', 'lat', and 'lon' must
    be present.
    
    Args:
        filepath: path to .xlsx file
        sheet_name: optional sheet name (defaults to first sheet)
    
    Returns:
        List of Waypoint objects, sorted by name number
    """
    try:
        import openpyxl
    except ImportError:
        raise ImportError(
            "Il pacchetto 'openpyxl' è necessario per leggere file XLSX. "
            "Installare con: pip install openpyxl"
        )
    
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    
    if sheet_name:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Foglio '{sheet_name}' non trovato. Fogli disponibili: {wb.sheetnames}")
        ws = wb[sheet_name]
    else:
        ws = wb.active
    
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    
    if not rows:
        raise ValueError("Il file XLSX è vuoto.")
    
    # First row = headers
    headers = [str(h).strip() if h else '' for h in rows[0]]
    col_map = _detect_columns(headers)
    
    # Validate required columns
    missing = []
    for required in ['name', 'lat', 'lon']:
        if required not in col_map:
            missing.append(required)
    
    if missing:
        raise ValueError(
            f"Colonne obbligatorie mancanti: {', '.join(missing)}. "
            f"Colonne trovate: {', '.join(headers)}. "
            f"Assicurarsi che il file contenga colonne per nome, latitudine e longitudine."
        )
    
    waypoints = []
    
    for row_idx, row in enumerate(rows[1:], start=2):
        if not row or all(cell is None for cell in row):
            continue  # Skip empty rows
        
        try:
            name_val = row[col_map['name']]
            lat_val = row[col_map['lat']]
            lon_val = row[col_map['lon']]
            
            if name_val is None or lat_val is None or lon_val is None:
                continue
            
            name = str(name_val).strip()
            lat = float(lat_val)
            lon = float(lon_val)
            
            # Validate coordinates
            if not (-90 <= lat <= 90) or not (-180 <= lon <= 180):
                print(f"[XLSX] Riga {row_idx}: coordinate non valide ({lat}, {lon}), skip")
                continue
            
            # Optional fields
            elevation = 0.0
            if 'elevation' in col_map and row[col_map['elevation']] is not None:
                try:
                    elevation = float(row[col_map['elevation']])
                except (ValueError, TypeError):
                    pass
            
            description = ""
            if 'description' in col_map and row[col_map['description']] is not None:
                description = str(row[col_map['description']]).strip()
            
            city = ""
            if 'city' in col_map and row[col_map['city']] is not None:
                city = str(row[col_map['city']]).strip()
            
            province = ""
            if 'province' in col_map and row[col_map['province']] is not None:
                province = str(row[col_map['province']]).strip()
            
            # Golden Point detection
            is_gp = False
            if 'golden' in col_map and row[col_map['golden']] is not None:
                gp_val = str(row[col_map['golden']]).strip().upper()
                is_gp = gp_val in ('GP', 'SI', 'SÌ', 'YES', 'TRUE', '1', 'X', 'GOLDEN')
            
            # Also check name for GP
            if not is_gp and 'GP' in name.upper().split():
                is_gp = True
            
            # Road classification from description
            road_type, is_unpaved, _ = classify_road_type(description)
            
            wp = Waypoint(
                id=len(waypoints),
                name=name,
                lat=lat,
                lon=lon,
                elevation=elevation,
                description=description,
                city=city,
                province=province,
                country="ITA",
                is_golden_point=is_gp,
                is_unpaved=is_unpaved,
                road_type=road_type,
                symbol="Flag, Red" if is_gp else "Flag, Blue",
            )
            waypoints.append(wp)
            
        except (ValueError, TypeError, IndexError) as e:
            print(f"[XLSX] Riga {row_idx}: errore parsing - {e}")
            continue
    
    if not waypoints:
        raise ValueError(
            "Nessun waypoint valido trovato nel file XLSX. "
            "Verificare che il file contenga righe con nome, latitudine e longitudine."
        )
    
    # Sort by numeric part of name
    def sort_key(wp):
        try:
            # Extract numeric prefix
            num = re.match(r'^(\d+)', wp.name)
            return int(num.group(1)) if num else 9999
        except:
            return 9999
    
    waypoints.sort(key=sort_key)
    
    # Re-assign sequential IDs after sorting
    for i, wp in enumerate(waypoints):
        wp.id = i
    
    print(f"[XLSX] Caricati {len(waypoints)} waypoints da {os.path.basename(filepath)}")
    return waypoints


def get_xlsx_sheet_names(filepath: str) -> list[str]:
    """Return list of sheet names from an XLSX file."""
    try:
        import openpyxl
    except ImportError:
        return []
    
    wb = openpyxl.load_workbook(filepath, read_only=True)
    names = wb.sheetnames
    wb.close()
    return names


if __name__ == "__main__":
    import sys
    if len(sys.argv) < 2:
        print("Uso: python xlsx_parser.py <file.xlsx>")
        sys.exit(1)
    
    wps = parse_xlsx(sys.argv[1])
    print(f"Parsed {len(wps)} waypoints")
    for wp in wps[:5]:
        print(f"  {wp.display_name} ({wp.lat:.4f}, {wp.lon:.4f}) elev={wp.elevation:.0f}m")
